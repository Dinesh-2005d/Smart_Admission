#!/usr/bin/env python3
"""
excel_reporter.py – Generates Automation_Test_Report.xlsx from pytest JSON report
Dark-themed professional Excel report for Smart Admission Selenium tests.
"""

import argparse
import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ─── Styles ───────────────────────────────────────────────────────────────────
DARK_BG = PatternFill("solid", fgColor="0F172A")
HEADER_BG = PatternFill("solid", fgColor="1E293B")
PASS_BG = PatternFill("solid", fgColor="14532D")
FAIL_BG = PatternFill("solid", fgColor="7F1D1D")
SKIP_BG = PatternFill("solid", fgColor="713F12")
ROW_ALT = PatternFill("solid", fgColor="1E293B")
ROW_EVEN = PatternFill("solid", fgColor="0F172A")

WHITE_FONT = Font(color="E2E8F0", size=10)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PASS_FONT = Font(color="86EFAC", size=10, bold=True)
FAIL_FONT = Font(color="FCA5A5", size=10, bold=True)
SKIP_FONT = Font(color="FDE68A", size=10, bold=True)
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
ACCENT_FONT = Font(bold=True, color="60A5FA", size=11)

THIN_BORDER = Border(
    left=Side(style="thin", color="334155"),
    right=Side(style="thin", color="334155"),
    top=Side(style="thin", color="334155"),
    bottom=Side(style="thin", color="334155"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def styled_cell(ws, row, col, value, font=WHITE_FONT, fill=DARK_BG, align=LEFT):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.fill = fill
    cell.alignment = align
    cell.border = THIN_BORDER
    return cell


def load_results(path):
    if not path or not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_test_results_sheet(wb, results, base_url):
    ws = wb.active
    ws.title = "Test Results"
    ws.sheet_view.showGridLines = False
    ws.tab_color = "2563EB"

    # Title
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "🧪 Smart Admission – Selenium E2E Test Report"
    title_cell.font = TITLE_FONT
    title_cell.fill = DARK_BG
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 36

    # Info row
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    summary = results.get("summary", {})
    ws.merge_cells("A2:H2")
    ws["A2"].value = f"URL: {base_url}  |  Date: {now}  |  Total: {summary.get('total',0)}  |  Passed: {summary.get('passed',0)}  |  Failed: {summary.get('failed',0)}"
    ws["A2"].font = Font(color="94A3B8", size=10)
    ws["A2"].fill = DARK_BG
    ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        ("#", 6), ("Test ID", 14), ("Test Name", 36), ("Module", 22),
        ("Status", 12), ("Duration (s)", 14), ("Failure Reason", 50), ("Screenshot", 20),
    ]
    for col, (title, width) in enumerate(headers, 1):
        styled_cell(ws, 3, col, title, font=HEADER_FONT, fill=HEADER_BG, align=CENTER)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 24

    # Test data rows
    tests = results.get("tests", [])
    for idx, test in enumerate(tests, 1):
        row = idx + 3
        ws.row_dimensions[row].height = 32

        nodeid = test.get("nodeid", "")
        parts = nodeid.split("::")
        module = parts[0].replace("tests/", "") if parts else ""
        test_name = parts[-1] if parts else nodeid
        outcome = test.get("outcome", "unknown")
        
        call_info = test.get("call", {})
        duration = call_info.get("duration", 0) if isinstance(call_info, dict) else 0
        reason = ""
        if outcome == "failed" and isinstance(call_info, dict):
            reason = str(call_info.get("longrepr", ""))[:200]

        # Status styling
        if outcome == "passed":
            status_font, status_fill = PASS_FONT, PASS_BG
            status_text = "✅ PASSED"
        elif outcome == "failed":
            status_font, status_fill = FAIL_FONT, FAIL_BG
            status_text = "❌ FAILED"
        elif outcome == "skipped":
            status_font, status_fill = SKIP_FONT, SKIP_BG
            status_text = "⏭️ SKIPPED"
        else:
            status_font, status_fill = WHITE_FONT, DARK_BG
            status_text = outcome.upper()

        row_fill = ROW_ALT if idx % 2 == 0 else ROW_EVEN

        styled_cell(ws, row, 1, idx, align=CENTER, fill=row_fill)
        styled_cell(ws, row, 2, f"TC-{idx:03d}", align=CENTER, fill=row_fill)
        styled_cell(ws, row, 3, test_name, fill=row_fill)
        styled_cell(ws, row, 4, module, fill=row_fill)
        styled_cell(ws, row, 5, status_text, font=status_font, fill=status_fill, align=CENTER)
        styled_cell(ws, row, 6, f"{duration:.2f}", align=CENTER, fill=row_fill)
        styled_cell(ws, row, 7, reason, fill=row_fill)
        styled_cell(ws, row, 8, f"FAIL_{test_name}.png" if outcome == "failed" else "—", fill=row_fill, align=CENTER)

    ws.freeze_panes = "A4"


def build_summary_sheet(wb, results):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "16A34A"

    summary = results.get("summary", {})
    total = summary.get("total", 0)
    passed = summary.get("passed", 0)
    failed = summary.get("failed", 0)
    skipped = summary.get("skipped", 0)
    duration = summary.get("duration", 0)
    pass_rate = f"{(passed / total * 100):.1f}%" if total > 0 else "N/A"

    ws.merge_cells("A1:D1")
    ws["A1"].value = "📊 Execution Summary"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = DARK_BG
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 36

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    data = [
        ("", "Metric", "Value", ""),
        ("📊", "Total Tests", str(total), ""),
        ("✅", "Passed", str(passed), pass_rate),
        ("❌", "Failed", str(failed), ""),
        ("⏭️", "Skipped", str(skipped), ""),
        ("⏱️", "Duration", f"{duration:.1f}s", ""),
        ("🎯", "Pass Rate", pass_rate, ""),
    ]

    for row_idx, (emoji, metric, value, extra) in enumerate(data, 2):
        ws.row_dimensions[row_idx].height = 28
        styled_cell(ws, row_idx, 1, emoji, align=CENTER)
        styled_cell(ws, row_idx, 2, metric, font=ACCENT_FONT if row_idx == 2 else WHITE_FONT)
        styled_cell(ws, row_idx, 3, value, align=CENTER,
                    font=PASS_FONT if "Pass" in metric else WHITE_FONT)
        styled_cell(ws, row_idx, 4, extra, align=CENTER)


def build_environment_sheet(wb, results, base_url):
    ws = wb.create_sheet("Environment")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "EAB308"

    ws.merge_cells("A1:C1")
    ws["A1"].value = "🌐 Test Environment"
    ws["A1"].font = TITLE_FONT
    ws["A1"].fill = DARK_BG
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 36

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20

    env_data = results.get("environment", {})

    rows = [
        ("Property", "Value", ""),
        ("Base URL", base_url, ""),
        ("Platform", env_data.get("Platform", "Linux"), ""),
        ("Python", env_data.get("Python", "3.11"), ""),
        ("Browser", "Chrome (Headless)", ""),
        ("Framework", "Selenium + Pytest", ""),
        ("Report Date", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"), ""),
    ]

    for row_idx, (prop, val, extra) in enumerate(rows, 2):
        ws.row_dimensions[row_idx].height = 26
        styled_cell(ws, row_idx, 1, prop, font=ACCENT_FONT if row_idx == 2 else WHITE_FONT)
        styled_cell(ws, row_idx, 2, val)
        styled_cell(ws, row_idx, 3, extra)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--results", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--base-url", default="")
    parser.add_argument("--platform", default="Web")
    parser.add_argument("--build-number", default="0")
    args = parser.parse_args()

    results = load_results(args.results)
    if results is None:
        print("⚠️  No results found; generating empty report")
        results = {
            "summary": {"total": 0, "passed": 0, "failed": 0, "skipped": 0, "duration": 0},
            "tests": [],
            "environment": {},
        }

    os.makedirs(os.path.dirname(args.output), exist_ok=True)

    wb = openpyxl.Workbook()
    build_test_results_sheet(wb, results, args.base_url)
    build_summary_sheet(wb, results)
    build_environment_sheet(wb, results, args.base_url)
    wb.save(args.output)

    total = results.get("summary", {}).get("total", 0)
    passed = results.get("summary", {}).get("passed", 0)
    print(f"✅ Excel report saved: {args.output}")
    print(f"   Tests: {total} | Passed: {passed}")


if __name__ == "__main__":
    main()
