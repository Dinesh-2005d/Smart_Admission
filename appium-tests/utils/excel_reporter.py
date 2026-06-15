#!/usr/bin/env python3
"""
excel_reporter.py – Generates Automation_Test_Report.xlsx for Appium tests
"""

import argparse
import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

DARK_BG = PatternFill("solid", fgColor="0F172A")
HEADER_BG = PatternFill("solid", fgColor="1E293B")
PASS_BG = PatternFill("solid", fgColor="14532D")
FAIL_BG = PatternFill("solid", fgColor="7F1D1D")
SKIP_BG = PatternFill("solid", fgColor="713F12")
ROW_ALT = PatternFill("solid", fgColor="1E293B")
ROW_EVEN = PatternFill("solid", fgColor="0F172A")

WHITE_FONT = Font(color="E2E8F0", size=10)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PASS_FONT = Font(color="86EFAC", size=10, bold=True)
FAIL_FONT = Font(color="FCA5A5", size=10, bold=True)
SKIP_FONT = Font(color="FDE68A", size=10, bold=True)
TITLE_FONT = Font(bold=True, color="FFFFFF", size=14)
ACCENT_FONT = Font(bold=True, color="60A5FA", size=11)

THIN_BORDER = Border(
    left=Side(style="thin", color="334155"),
    right=Side(style="thin", color="334155"),
    top=Side(style="thin", color="334155"),
    bottom=Side(style="thin", color="334155"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def styled_cell(ws, row, col, value, font=WHITE_FONT, fill=DARK_BG, align=LEFT):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font, cell.fill, cell.alignment, cell.border = font, fill, align, THIN_BORDER
    return cell


def load_results(path):
    if not path or not os.path.exists(path):
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_report(wb, results, platform, build_number):
    ws = wb.active
    ws.title = "Test Results"
    ws.sheet_view.showGridLines = False
    ws.tab_color = "2563EB"

    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    summary = results.get("summary", {})

    ws.merge_cells("A1:H1")
    ws["A1"].value = f"🤖 Smart Admission – {platform} Appium E2E Report"
    ws["A1"].font, ws["A1"].fill, ws["A1"].alignment = TITLE_FONT, DARK_BG, CENTER
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    ws["A2"].value = f"Build #{build_number}  |  {now}  |  Total: {summary.get('total',0)}  |  Passed: {summary.get('passed',0)}  |  Failed: {summary.get('failed',0)}"
    ws["A2"].font, ws["A2"].fill, ws["A2"].alignment = Font(color="94A3B8", size=10), DARK_BG, CENTER
    ws.row_dimensions[2].height = 20

    headers = [
        ("#", 6), ("Test ID", 14), ("Test Name", 36), ("Module", 22),
        ("Status", 12), ("Duration (s)", 14), ("Failure Reason", 50), ("Screenshot", 20),
    ]
    for col, (title, width) in enumerate(headers, 1):
        styled_cell(ws, 3, col, title, font=HEADER_FONT, fill=HEADER_BG, align=CENTER)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 24

    tests = results.get("tests", [])
    for idx, test in enumerate(tests, 1):
        row = idx + 3
        ws.row_dimensions[row].height = 32
        nodeid = test.get("nodeid", "")
        parts = nodeid.split("::")
        module = parts[0].replace("tests/", "") if parts else ""
        test_name = parts[-1] if parts else nodeid
        outcome = test.get("outcome", "unknown")
        call_info = test.get("call", {})
        duration = call_info.get("duration", 0) if isinstance(call_info, dict) else 0
        reason = ""
        if outcome == "failed" and isinstance(call_info, dict):
            reason = str(call_info.get("longrepr", ""))[:200]

        if outcome == "passed":
            sf, sfill, stxt = PASS_FONT, PASS_BG, "✅ PASSED"
        elif outcome == "failed":
            sf, sfill, stxt = FAIL_FONT, FAIL_BG, "❌ FAILED"
        elif outcome == "skipped":
            sf, sfill, stxt = SKIP_FONT, SKIP_BG, "⏭️ SKIPPED"
        else:
            sf, sfill, stxt = WHITE_FONT, DARK_BG, outcome.upper()

        rfill = ROW_ALT if idx % 2 == 0 else ROW_EVEN
        styled_cell(ws, row, 1, idx, align=CENTER, fill=rfill)
        styled_cell(ws, row, 2, f"TC-{idx:03d}", align=CENTER, fill=rfill)
        styled_cell(ws, row, 3, test_name, fill=rfill)
        styled_cell(ws, row, 4, module, fill=rfill)
        styled_cell(ws, row, 5, stxt, font=sf, fill=sfill, align=CENTER)
        styled_cell(ws, row, 6, f"{duration:.2f}", align=CENTER, fill=rfill)
        styled_cell(ws, row, 7, reason, fill=rfill)
        styled_cell(ws, row, 8, f"FAIL_{test_name}.png" if outcome == "failed" else "—", fill=rfill, align=CENTER)

    ws.freeze_panes = "A4"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.tab_color = "16A34A"
    total = summary.get("total", 0)
    passed = summary.get("passed", 0)
    failed = summary.get("failed", 0)
    skipped = summary.get("skipped", 0)
    dur = summary.get("duration", 0)
    pr = f"{(passed/total*100):.1f}%" if total > 0 else "N/A"

    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "📊 Execution Summary"
    ws2["A1"].font, ws2["A1"].fill, ws2["A1"].alignment = TITLE_FONT, DARK_BG, CENTER
    ws2.row_dimensions[1].height = 36
    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18

    data = [
        ("", "Metric", "Value", ""),
        ("📊", "Total Tests", str(total), ""),
        ("✅", "Passed", str(passed), pr),
        ("❌", "Failed", str(failed), ""),
        ("⏭️", "Skipped", str(skipped), ""),
        ("⏱️", "Duration", f"{dur:.1f}s", ""),
        ("🎯", "Pass Rate", pr, ""),
        ("📱", "Platform", platform, ""),
        ("🔢", "Build", f"#{build_number}", ""),
    ]
    for ri, (e, m, v, x) in enumerate(data, 2):
        ws2.row_dimensions[ri].height = 28
        styled_cell(ws2, ri, 1, e, align=CENTER)
        styled_cell(ws2, ri, 2, m, font=ACCENT_FONT if ri == 2 else WHITE_FONT)
        styled_cell(ws2, ri, 3, v, align=CENTER)
        styled_cell(ws2, ri, 4, x, align=CENTER)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--results", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--platform", default="Android")
    parser.add_argument("--build-number", default="0")
    parser.add_argument("--base-url", default="")
    args = parser.parse_args()

    results = load_results(args.results)
    if results is None:
        results = {"summary": {"total": 0, "passed": 0, "failed": 0, "skipped": 0, "duration": 0}, "tests": [], "environment": {}}

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    wb = openpyxl.Workbook()
    build_report(wb, results, args.platform, args.build_number)
    wb.save(args.output)
    print(f"✅ Excel: {args.output}")


if __name__ == "__main__":
    main()
