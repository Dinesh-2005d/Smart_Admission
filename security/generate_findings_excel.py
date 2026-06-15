#!/usr/bin/env python3
"""
Excel Report Generator for Smart Admission Security Findings
Generates findings.xlsx and endpoint-inventory.xlsx
"""

import argparse
import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ─── Colour palette ──────────────────────────────────────────────────────────
COLOURS = {
    "header_bg":   "0F172A",  # dark navy
    "header_fg":   "FFFFFF",
    "critical_bg": "7F1D1D",
    "critical_fg": "FCA5A5",
    "high_bg":     "7C2D12",
    "high_fg":     "FDBA74",
    "medium_bg":   "713F12",
    "medium_fg":   "FDE68A",
    "low_bg":      "14532D",
    "low_fg":      "86EFAC",
    "row_alt":     "1E293B",
    "row_even":    "0F172A",
    "text":        "E2E8F0",
    "accent":      "2563EB",
}


def thin_border():
    side = Side(style="thin", color="334155")
    return Border(left=side, right=side, top=side, bottom=side)


def header_style(ws, row, col, value, width=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True, color=COLOURS["header_fg"], size=11)
    cell.fill = PatternFill("solid", fgColor=COLOURS["header_bg"])
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def data_cell(ws, row, col, value, severity=None, center=False):
    cell = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    bg = COLOURS["row_alt"] if row % 2 == 0 else COLOURS["row_even"]
    fg = COLOURS["text"]

    if severity:
        sev_lower = severity.lower()
        if sev_lower == "critical":
            bg, fg = COLOURS["critical_bg"], COLOURS["critical_fg"]
        elif sev_lower == "high":
            bg, fg = COLOURS["high_bg"], COLOURS["high_fg"]
        elif sev_lower == "medium":
            bg, fg = COLOURS["medium_bg"], COLOURS["medium_fg"]
        elif sev_lower == "low":
            bg, fg = COLOURS["low_bg"], COLOURS["low_fg"]

    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(color=fg, size=10)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )
    cell.border = thin_border()
    return cell


def load_json(path):
    if not path or not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def collect_findings(semgrep_path, gitleaks_path, npm_path, trivy_path):
    all_findings = []

    # Semgrep
    sdata = load_json(semgrep_path)
    for r in sdata.get("results", []):
        sev_map = {"ERROR": "High", "WARNING": "Medium", "INFO": "Low"}
        sev_raw = r.get("extra", {}).get("severity", "INFO").upper()
        sev = sev_map.get(sev_raw, "Low")
        all_findings.append({
            "id": f"SG-{len(all_findings)+1:03d}",
            "tool": "Semgrep",
            "severity": sev,
            "type": "SAST",
            "rule": r.get("check_id", "")[:80],
            "file": r.get("path", ""),
            "line": str(r.get("start", {}).get("line", "")),
            "description": r.get("extra", {}).get("message", "")[:200],
            "fix": r.get("extra", {}).get("fix", "Review and remediate")[:200],
            "cvss": "",
            "cve": "",
        })

    # Gitleaks
    gdata = load_json(gitleaks_path)
    leaks = gdata.get("findings", gdata) if isinstance(gdata, dict) else gdata
    if isinstance(leaks, list):
        for leak in leaks:
            all_findings.append({
                "id": f"GL-{len(all_findings)+1:03d}",
                "tool": "Gitleaks",
                "severity": "High",
                "type": "Secret Detected",
                "rule": leak.get("RuleID", leak.get("rule", "generic-secret")),
                "file": leak.get("File", leak.get("file", "")),
                "line": str(leak.get("StartLine", leak.get("line", ""))),
                "description": leak.get("Description", "Potential secret/credential exposed"),
                "fix": "Remove secret and rotate immediately. Use GitHub Secrets instead.",
                "cvss": "7.5",
                "cve": "",
            })

    # Check proxy-server.js manually
    if os.path.exists("proxy-server.js"):
        with open("proxy-server.js") as f:
            content = f.read()
        if "Bearer" in content and "process.env" not in content:
            all_findings.append({
                "id": f"MS-{len(all_findings)+1:03d}",
                "tool": "Manual Scan",
                "severity": "High",
                "type": "Hardcoded Secret",
                "rule": "hardcoded-api-key",
                "file": "proxy-server.js",
                "line": "20",
                "description": "Hardcoded Groq API Bearer token found in source code",
                "fix": "Use process.env.GROQ_API_KEY. Store key in GitHub Secrets.",
                "cvss": "7.5",
                "cve": "CWE-798",
            })

    # npm audit
    ndata = load_json(npm_path)
    for pkg, info in ndata.get("vulnerabilities", {}).items():
        sev = info.get("severity", "low").capitalize()
        via = info.get("via", [])
        desc = ""
        cve = ""
        if via and isinstance(via[0], dict):
            desc = via[0].get("title", f"Vulnerability in {pkg}")
            cve = via[0].get("url", "")
        else:
            desc = f"Vulnerable dependency: {pkg}"
        all_findings.append({
            "id": f"NPM-{len(all_findings)+1:03d}",
            "tool": "npm audit",
            "severity": sev,
            "type": "Dependency",
            "rule": f"vuln-{pkg}",
            "file": "package.json",
            "line": "N/A",
            "description": desc[:200],
            "fix": f"Run `npm audit fix` or upgrade {pkg}",
            "cvss": "",
            "cve": cve,
        })

    # Trivy
    tdata = load_json(trivy_path)
    for result in tdata.get("Results", []):
        for vuln in result.get("Vulnerabilities", []):
            sev = vuln.get("Severity", "LOW").capitalize()
            all_findings.append({
                "id": f"TV-{len(all_findings)+1:03d}",
                "tool": "Trivy",
                "severity": sev,
                "type": "CVE",
                "rule": vuln.get("VulnerabilityID", ""),
                "file": result.get("Target", ""),
                "line": "N/A",
                "description": (vuln.get("Description") or vuln.get("Title", ""))[:200],
                "fix": f"Update {vuln.get('PkgName','')} to {vuln.get('FixedVersion','latest')}",
                "cvss": str(vuln.get("CVSS", {}).get("nvd", {}).get("V3Score", "")),
                "cve": vuln.get("VulnerabilityID", ""),
            })

    return all_findings


def build_findings_sheet(wb, findings):
    ws = wb.active
    ws.title = "Security Findings"
    ws.sheet_view.showGridLines = False
    ws.tab_color = "2563EB"

    # Title row
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "🔐 Smart Admission — Security Findings Report"
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="0F172A")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Sub-header info
    ws.merge_cells("A2:L2")
    ws["A2"].value = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | Total Findings: {len(findings)}"
    ws["A2"].font = Font(size=10, color="94A3B8")
    ws["A2"].fill = PatternFill("solid", fgColor="0F172A")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Headers
    headers = [
        ("ID", 8), ("Tool", 14), ("Severity", 12), ("Type", 18),
        ("Rule / CVE", 30), ("File", 32), ("Line", 8),
        ("Description", 50), ("Recommended Fix", 40),
        ("CVSS", 8), ("CVE", 18), ("Status", 12),
    ]
    ws.row_dimensions[3].height = 22
    for col, (title, width) in enumerate(headers, 1):
        header_style(ws, 3, col, title, width)

    # Data rows
    severity_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    sorted_f = sorted(findings, key=lambda x: severity_order.get(x.get("severity","Low"), 4))

    for row_idx, f in enumerate(sorted_f, 4):
        sev = f.get("severity", "Low")
        ws.row_dimensions[row_idx].height = 36
        values = [
            f.get("id", ""),
            f.get("tool", ""),
            sev,
            f.get("type", ""),
            f.get("rule", ""),
            f.get("file", ""),
            f.get("line", ""),
            f.get("description", ""),
            f.get("fix", ""),
            f.get("cvss", ""),
            f.get("cve", ""),
            "Open",
        ]
        for col_idx, val in enumerate(values, 1):
            data_cell(ws, row_idx, col_idx, val,
                      severity=sev if col_idx == 3 else None,
                      center=(col_idx in (1, 3, 6, 7, 10, 12)))

    # Freeze panes
    ws.freeze_panes = "A4"


def build_endpoint_inventory(wb):
    """Build the App Endpoint Inventory sheet."""
    ws = wb.create_sheet("Endpoint Inventory")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "16A34A"

    ws.merge_cells("A1:F1")
    ws["A1"].value = "📡 Smart Admission — App Screen / Endpoint Inventory"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0F172A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        ("Screen / Endpoint", 36), ("Type", 14), ("Auth Required", 15),
        ("Role", 14), ("File Path", 42), ("Notes", 36),
    ]
    for col, (title, width) in enumerate(headers, 1):
        header_style(ws, 2, col, title, width)

    endpoints = [
        ("/ (Splash)", "Screen", "No", "Public", "src/screens/AnimatedSplashScreen.js", "Entry point"),
        ("/Home", "Screen", "No", "Public", "src/screens/HomeScreen.js", "State & Board selection"),
        ("/MarksEntry", "Screen", "No", "Public", "src/screens/MarksEntryScreen.js", "Marks & dept entry"),
        ("/CollegeList", "Screen", "No", "Public", "src/screens/CollegeListScreen.js", "AI-filtered results"),
        ("/Details", "Screen", "No", "Public", "src/screens/DetailsScreen.js", "College details"),
        ("/Search", "Screen", "No", "Public", "src/screens/SearchScreen.js", "College search"),
        ("/Compare", "Screen", "No", "Public", "src/screens/CompareScreen.js", "Side-by-side compare"),
        ("/CollegeChat", "Screen", "No", "Public", "src/screens/CollegeChatScreen.js", "AI chat about college"),
        ("POST /claude", "API Proxy", "No", "Public", "proxy-server.js", "⚠️ Hardcoded API key"),
        ("Groq API /openai/v1/chat/completions", "External API", "Yes (key)", "Service", "proxy-server.js", "Third-party LLM"),
    ]

    for row_idx, ep in enumerate(endpoints, 3):
        ws.row_dimensions[row_idx].height = 28
        for col_idx, val in enumerate(ep, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            bg = "1E293B" if row_idx % 2 == 0 else "0F172A"
            fg = "FCA5A5" if "⚠️" in str(val) else "E2E8F0"
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(color=fg, size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border()

    ws.freeze_panes = "A3"


def build_dependency_sheet(wb, findings):
    ws = wb.create_sheet("Dependency Vulnerabilities")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "EAB308"

    ws.merge_cells("A1:G1")
    ws["A1"].value = "📦 Dependency Vulnerabilities"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0F172A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        ("ID", 10), ("Tool", 12), ("Severity", 12),
        ("Package / Rule", 28), ("Description", 50), ("Fix", 40), ("CVE", 20),
    ]
    for col, (title, width) in enumerate(headers, 1):
        header_style(ws, 2, col, title, width)

    dep_findings = [f for f in findings if f.get("type") in ("Dependency", "CVE")]

    for row_idx, f in enumerate(dep_findings, 3):
        sev = f.get("severity", "Low")
        ws.row_dimensions[row_idx].height = 32
        values = [
            f.get("id", ""), f.get("tool", ""), sev,
            f.get("rule", ""), f.get("description", ""),
            f.get("fix", ""), f.get("cve", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            data_cell(ws, row_idx, col_idx, val,
                      severity=sev if col_idx == 3 else None,
                      center=(col_idx in (1, 2, 3)))

    ws.freeze_panes = "A3"


def build_risk_summary(wb, findings):
    ws = wb.create_sheet("Risk Summary")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "EF4444"

    counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
    for f in findings:
        sev = f.get("severity", "Low")
        if sev in counts:
            counts[sev] += 1

    total = sum(counts.values())
    score = max(0, 100 - counts["Critical"]*15 - counts["High"]*8 - counts["Medium"]*3 - counts["Low"])

    ws.merge_cells("A1:D1")
    ws["A1"].value = "🎯 Risk Summary Dashboard"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0F172A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    summary_data = [
        ("", "Severity", "Count", "% of Total"),
        ("🚨", "Critical", counts["Critical"], f"{counts['Critical']/max(total,1)*100:.1f}%"),
        ("🔴", "High",     counts["High"],     f"{counts['High']/max(total,1)*100:.1f}%"),
        ("🟡", "Medium",   counts["Medium"],   f"{counts['Medium']/max(total,1)*100:.1f}%"),
        ("🟢", "Low",      counts["Low"],      f"{counts['Low']/max(total,1)*100:.1f}%"),
        ("",   "Total",    total,              "100%"),
        ("",   "Security Score", f"{score}/100", ""),
    ]

    col_widths = [6, 20, 14, 16]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    for row_idx, row_data in enumerate(summary_data, 2):
        sev_name = row_data[1] if len(row_data) > 1 else ""
        sev_for_color = sev_name if sev_name in counts else None
        ws.row_dimensions[row_idx].height = 26
        for col_idx, val in enumerate(row_data, 1):
            data_cell(ws, row_idx, col_idx, val,
                      severity=sev_for_color if col_idx in (1, 2) else None,
                      center=True)

    ws.freeze_panes = "A2"


def main():
    parser = argparse.ArgumentParser(description="Generate security Excel reports")
    parser.add_argument("--semgrep", default="")
    parser.add_argument("--gitleaks", default="")
    parser.add_argument("--npm-audit", default="")
    parser.add_argument("--trivy", default="")
    parser.add_argument("--findings-output", required=True)
    parser.add_argument("--endpoints-output", required=True)
    args = parser.parse_args()

    findings = collect_findings(
        args.semgrep, args.gitleaks, args.npm_audit, args.trivy
    )

    print(f"📊 Total findings collected: {len(findings)}")

    # ── findings.xlsx ─────────────────────────────────────────────────────────
    os.makedirs(os.path.dirname(args.findings_output), exist_ok=True)
    wb_findings = openpyxl.Workbook()
    build_findings_sheet(wb_findings, findings)
    build_dependency_sheet(wb_findings, findings)
    build_risk_summary(wb_findings, findings)
    wb_findings.save(args.findings_output)
    print(f"✅ findings.xlsx → {args.findings_output}")

    # ── endpoint-inventory.xlsx ───────────────────────────────────────────────
    os.makedirs(os.path.dirname(args.endpoints_output), exist_ok=True)
    wb_ep = openpyxl.Workbook()
    # Remove default sheet
    default = wb_ep.active
    build_endpoint_inventory(wb_ep)
    wb_ep.remove(default)
    wb_ep.save(args.endpoints_output)
    print(f"✅ endpoint-inventory.xlsx → {args.endpoints_output}")


if __name__ == "__main__":
    main()
